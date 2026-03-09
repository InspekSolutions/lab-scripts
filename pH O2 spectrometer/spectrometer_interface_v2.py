# This script is used to control the OceanMini spectrometer
# It is used to capture spectra of the pH probe
# It is used to save the spectra to a CSV file

from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar

import sys, time, traceback, threading
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl import load_workbook, Workbook
import os
import re
import pandas as pd
from scipy.optimize import curve_fit
from scipy.signal import savgol_filter, find_peaks, peak_widths

import seabreeze
seabreeze.use('pyseabreeze')
import seabreeze.spectrometers as sb

from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


# ==================== pH CALIBRATION CONSTANTS & FUNCTIONS ====================
WL_RANGE = (500, 700)               # nm, wavelength range for fitting
SMOOTH_POLY = 2                     # Savitzky-Golay polynomial order
SMOOTH_MIN_WIN = 21                 # odd window for smoothing
SIGMA_MIN, SIGMA_MAX = 2.0, 80.0    # nm, Gaussian width bounds
MAXFEV = 60000                      # maximum function evaluations for fitting

def extract_ph(s):
    """Extract pH value from column header like 'pH 7.5'"""
    m = re.search(r'pH\s*([0-9]+[.,]?[0-9]*)', str(s), re.IGNORECASE)
    return np.nan if not m else float(m.group(1).replace(',', '.'))

def to_float(s):
    """Convert pandas series to float, handling commas as decimal separators"""
    return pd.to_numeric(s.astype(str).str.replace(',', '.', regex=False), errors='coerce')

def parse_xy_pairs(df):
    """Parse Excel dataframe into wavelength-intensity pairs with pH values"""
    pairs = []
    for i in range(0, df.shape[1], 2):
        if i+1 >= df.shape[1]: break
        xname, yname = df.columns[i], df.columns[i+1]
        pH = extract_ph(xname)
        X = to_float(df.iloc[:, i]).values
        Y = to_float(df.iloc[:, i+1]).values
        m = np.isfinite(X) & np.isfinite(Y)
        if m.sum() < 9 or np.isnan(pH): 
            continue
        x, y = X[m], Y[m]
        order = np.argsort(x)
        x, y = x[order], y[order]
        if WL_RANGE:
            lo, hi = WL_RANGE
            mm = (x >= lo) & (x <= hi)
            x, y = x[mm], y[mm]
        if len(x) < 9: 
            continue
        pairs.append({"pH": float(pH), "x": x, "y": y, "label": f'{yname} vs {xname}'})
    return pairs

def auto_win(n):
    """Automatically determine Savitzky-Golay window size"""
    if n < 7: return 5 if n >= 5 else 3
    w = SMOOTH_MIN_WIN if n >= SMOOTH_MIN_WIN else (n//2)*2 + 1
    if w > n: w = n if n % 2 == 1 else n-1
    return max(3, w)

def double_gaussian(x, A1, x01, s1, A2, x02, s2, c):
    """Double Gaussian model"""
    g1 = A1 * np.exp(-0.5*((x-x01)/s1)**2)
    g2 = A2 * np.exp(-0.5*((x-x02)/s2)**2)
    return g1 + g2 + c

def seed_from_data(x, y):
    """Generate initial guess for double Gaussian fit"""
    pk, _ = find_peaks(y)
    i0 = pk[np.argmax(y[pk])] if len(pk) else int(np.argmax(y))
    A1, x01 = float(y[i0]), float(x[i0])
    try:
        w = peak_widths(y, [i0], rel_height=0.5)[0][0] * np.mean(np.diff(x))
        s_est = float(np.clip(w/2.355, SIGMA_MIN, SIGMA_MAX))
    except Exception:
        s_est = 12.0
    thr = A1 * 0.06
    pk2, _ = find_peaks(y, prominence=thr)
    pk2 = [i for i in pk2 if abs(i - i0) > 3]
    if len(pk2) > 0:
        j = pk2[np.argmax(y[pk2])]
        A2, x02 = float(y[j]), float(x[j])
    else:
        A2, x02 = 0.5*A1, float(min(x.max(), x01 + 18))
    if x02 < x01:
        (A1,x01), (A2,x02) = (A2,x02), (A1,x01)
    c0 = float(np.percentile(y, 5))
    return [A1, x01, s_est, A2, x02, s_est, c0]

def fit_left_center_lambda(x, y):
    """Fit double Gaussian and return the left peak center wavelength"""
    # Filter data to only include wavelengths >= 480 nm
    mask = x >= 480.0
    if not np.any(mask):
        # If no data above 480 nm, use all data
        x_filt, y_filt = x, y
    else:
        x_filt, y_filt = x[mask], y[mask]
    
    Y = savgol_filter(y_filt, auto_win(len(y_filt)), SMOOTH_POLY) if len(y_filt) >= 7 else y_filt
    p0 = seed_from_data(x_filt, Y)
    lb = [0.0, x_filt.min(), SIGMA_MIN, 0.0, x_filt.min(), SIGMA_MIN, -np.inf]
    ub = [np.inf, x_filt.max(), SIGMA_MAX, np.inf, x_filt.max(), SIGMA_MAX,  np.inf]
    popt, _ = curve_fit(double_gaussian, x_filt, Y, p0=p0, bounds=(lb, ub),
                        method='trf', maxfev=MAXFEV, ftol=1e-10, xtol=1e-10, loss='soft_l1')
    A1,x01,s1, A2,x02,s2, c = popt
    left_x0 = x01 if x01 < x02 else x02
    return float(left_x0)

def I_model_pH(pH, Ia, Ib, Ka):
    """Henderson-Hasselbalch-like intensity model"""
    H = 10.0**(-pH)
    return Ia*(H/(Ka+H)) + Ib*(Ka/(Ka+H))

def fit_Ia_Ib_Ka(pH_arr, I_arr):
    """Fit calibration constants from pH and intensity arrays"""
    Ia0, Ib0 = float(I_arr[np.argmin(pH_arr)]), float(I_arr[np.argmax(pH_arr)])
    Ka0 = 10.0**(-np.median(pH_arr))
    popt, pcov = curve_fit(I_model_pH, pH_arr, I_arr, p0=[Ia0, Ib0, Ka0],
                           bounds=([0,0,1e-12],[np.inf,np.inf,1.0]), method='trf', maxfev=40000)
    Ia, Ib, Ka = popt
    return Ia, Ib, Ka

def invert_to_pH(I, Ia, Ib, Ka):
    """Invert the model to predict pH from intensity"""
    H = Ka * (I - Ib) / (Ia - I)
    H = np.where(np.isfinite(H), H, np.nan)
    H = np.clip(H, 1e-12, 1.0, where=np.isfinite(H))
    return -np.log10(H)

# ==============================================================================


class SpectrumWorker(QtCore.QObject):
    spectrumReady = QtCore.pyqtSignal(np.ndarray, np.ndarray, float, dict)
    deviceStatus = QtCore.pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._device = None
        self._running = False
        self._lock = threading.Lock()       # params
        self._io_lock = threading.RLock()   # **serialize all device I/O**
        self._integration_time_us = 200_000
        self._correct_edark = False
        self._correct_nonlinearity = False
        self._apply_dark_sub = True
        self._dark = None
        self._avg_window = 1
        self._pause_once = False            # request one-frame pause for critical ops

    # ---- device controls ----
    @QtCore.pyqtSlot()
    def connect_device(self):
        try:
            with self._io_lock:
                devs = sb.list_devices()
                if not devs:
                    self.deviceStatus.emit("No devices found")
                    return False
                # avoid reopening if already open
                if self._device is None:
                    self._device = sb.Spectrometer.from_first_available()
                    self._device.integration_time_micros(self._integration_time_us)
                self.deviceStatus.emit(f"Connected, integration time: {self._integration_time_us}")
                return True
        except Exception as e:
            self._device = None
            self.deviceStatus.emit(f"Connect error: {e}")
            return False

    @QtCore.pyqtSlot()
    def disconnect_device(self):
        try:
            with self._io_lock:
                if self._device is not None:
                    self._device.close()
        except Exception:
            pass
        finally:
            self._device = None
            self.deviceStatus.emit("Disconnected")

    # ---- setters (all safe with I/O lock) ----
    @QtCore.pyqtSlot(int)
    def set_integration_ms(self, ms: int):
        us = max(10, int(ms) * 1000)
        with self._lock:
            self._integration_time_us = us
        try:
            with self._io_lock:
                if self._device is not None:
                    self._device.integration_time_micros(us)
        except Exception as e:
            self.deviceStatus.emit(f"Set IT error: {e}")

    @QtCore.pyqtSlot(bool)
    def set_correct_edark(self, enabled: bool):
        with self._lock: self._correct_edark = bool(enabled)

    @QtCore.pyqtSlot(bool)
    def set_correct_nonlinearity(self, enabled: bool):
        with self._lock: self._correct_nonlinearity = bool(enabled)

    @QtCore.pyqtSlot(bool)
    def set_apply_dark_sub(self, enabled: bool):
        with self._lock: self._apply_dark_sub = bool(enabled)

    @QtCore.pyqtSlot(int)
    def set_avg_window(self, n: int):
        with self._lock: self._avg_window = max(1, int(n))

    @QtCore.pyqtSlot()
    def capture_dark(self):
        """Capture current spectrum as dark, safely."""
        if self._device is None:
            self.deviceStatus.emit("No device for dark capture")
            return
        try:
            # request the loop to skip one acquisition window
            self._pause_once = True
            with self._io_lock:
                # read once synchronously
                intens = self._device.intensities(
                    correct_dark_counts=self._correct_edark,
                    correct_nonlinearity=self._correct_nonlinearity
                )
            with self._lock:
                self._dark = intens.copy()
            self.deviceStatus.emit("Dark captured")
        except Exception as e:
            self.deviceStatus.emit(f"Dark capture error: {e}")

    @QtCore.pyqtSlot()
    def clear_dark(self):
        with self._lock:
            self._dark = None
        self.deviceStatus.emit("Dark cleared")

    # ---- run/stop streaming ----
    @QtCore.pyqtSlot()
    def start(self):
        if self._running:
            return
        self._running = True
        if self._device is None:
            self.connect_device()
        threading.Thread(target=self._loop, daemon=True).start()
        self.deviceStatus.emit("Streaming started")

    @QtCore.pyqtSlot()
    def stop(self):
        self._running = False
        self.deviceStatus.emit("Streaming stopped")

    # ---- acquisition loop (all device calls under _io_lock) ----
    def _loop(self):
        wl = None
        while self._running:
            try:
                if self._device is None:
                    if not self.connect_device():
                        time.sleep(0.25)
                        continue

                if self._pause_once:
                    # give room for a control action (e.g., capture_dark)
                    self._pause_once = False
                    time.sleep(0.01)

                with self._lock:
                    correct_ed = self._correct_edark
                    correct_nl = self._correct_nonlinearity
                    apply_dark = self._apply_dark_sub
                    dark = None if self._dark is None else self._dark.copy()
                    avgN = self._avg_window
                    it_us = self._integration_time_us

                with self._io_lock:
                    if wl is None:
                        wl = self._device.wavelengths()
                    if avgN <= 1:
                        intens = self._device.intensities(
                            correct_dark_counts=correct_ed,
                            correct_nonlinearity=correct_nl
                        )
                    else:
                        acc = None
                        for _ in range(avgN):
                            arr = self._device.intensities(
                                correct_dark_counts=correct_ed,
                                correct_nonlinearity=correct_nl
                            )
                            if acc is None: acc = arr.astype(np.float64)
                            else: acc += arr
                        intens = (acc / float(avgN)).astype(np.float64)

                if apply_dark and dark is not None and dark.shape == intens.shape:
                    intens = intens - dark
                    intens[intens < 0] = 0

                ts = time.time()
                meta = {
                    "integration_us": it_us,
                    "correct_edark": correct_ed,
                    "correct_nonlinearity": correct_nl,
                    "apply_dark_sub": apply_dark,
                    "averages": avgN,
                }
                self.spectrumReady.emit(wl, intens, ts, meta)

            except Exception as e:
                # Only drop device on real I/O failures, not every exception
                self.deviceStatus.emit(f"I/O error, retrying: {e}")
                self.disconnect_device()
                time.sleep(0.1)

        # cleanup
        self.disconnect_device()


# --- replace the whole MplCanvas class with this ---
class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, max_fps=20):
        self.fig = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setParent(parent)

        self.ax.set_xlabel("Wavelength (nm)")
        self.ax.set_ylabel("Intensity (a.u.)")
        self._live_line = None           # single live trace
        self._x_fixed = False
        self._last_draw_t = 0.0
        self._frame_interval = 1.0 / float(max_fps)
        self._autoscale_y = True
        self._manual_y = False
        self._ymin = None
        self._ymax = None
        self._ymax_callback = None  # Callback to update spinbox

        # snapshots (overlaid)
        self._snap_lines = []            # list[Line2D]
        
        # Connect scroll event
        self.mpl_connect('scroll_event', self.on_scroll)

    def set_autoscale_y(self, enabled: bool):
        self._autoscale_y = bool(enabled)
        if enabled:
            self._manual_y = False
            # let next plot decide limits automatically
        self.draw_idle()

    def set_manual_y(self, enabled: bool, ymin: float = None, ymax: float = None):
        """Enable/disable manual Y and set limits."""
        self._manual_y = bool(enabled)
        if enabled:
            if ymin is None or ymax is None:
                return
            if ymax <= ymin:
                ymin, ymax = float(ymin), float(ymax)
                if ymax <= ymin:
                    ymax = ymin + 1.0  # fallback
            self._ymin, self._ymax = float(ymin), float(ymax)
            self.ax.set_ylim(self._ymin, self._ymax)
        self.draw_idle()

    def update_manual_y_limits(self, ymin: float, ymax: float):
        """Change limits while staying in manual mode."""
        if ymax <= ymin:
            ymax = ymin + 1.0
        self._ymin, self._ymax = float(ymin), float(ymax)
        self._manual_y = True
        self.ax.set_ylim(self._ymin, self._ymax)
        self.draw_idle()

    def reset_y(self):
        """Disable manual and autoscale once on next frame."""
        self._manual_y = False
        self._autoscale_y = True
        self.draw_idle()

    def plot_live(self, wl, inten):
        # Fix x once (wavelengths constant)
        if not self._x_fixed:
            self.ax.set_xlim(float(wl[0]), float(wl[-1]))
            self._x_fixed = True

        if self._live_line is None:
            (self._live_line,) = self.ax.plot(wl, inten, lw=1, label="Live")
        else:
            # only y changes
            self._live_line.set_ydata(inten)

        # Only touch Y limits if autoscale enabled (so zoom/pan isn’t overridden)
        if self._autoscale_y:
            ymin = 0.0 if np.amin(inten) >= 0 else float(np.amin(inten))
            ymax = float(np.amax(inten))
            if ymax <= ymin: ymax = ymin + 1.0
            margin = 0.05 * (ymax - ymin)
            self.ax.set_ylim(ymin - margin, ymax + margin)
            
        if self._manual_y:
            # keep fixed manual limits
            self.ax.set_ylim(self._ymin, self._ymax)
        elif self._autoscale_y:
            ymin = 0.0 if np.amin(inten) >= 0 else float(np.amin(inten))
            ymax = float(np.amax(inten))
            if ymax <= ymin: ymax = ymin + 1.0
            margin = 0.05 * (ymax - ymin)
            self.ax.set_ylim(ymin - margin, ymax + margin)

        now = time.time()
        if (now - self._last_draw_t) >= self._frame_interval:
            self.draw_idle()
            self._last_draw_t = now

    def add_snapshot(self, wl, inten, label: str):
        """Overlay a saved spectrum."""
        (line,) = self.ax.plot(wl, inten, lw=1, alpha=0.9, label=label)
        self._snap_lines.append(line)
        self._refresh_legend()
        self.draw_idle()

    def clear_snapshots(self):
        for ln in self._snap_lines:
            try:
                ln.remove()
            except Exception:
                pass
        self._snap_lines.clear()
        self._refresh_legend()
        self.draw_idle()

    def _refresh_legend(self):
        # Show legend only if we have at least one snapshot
        handles, labels = self.ax.get_legend_handles_labels()
        if len(labels) > 1:
            self.ax.legend(loc="best", frameon=False)
        else:
            leg = self.ax.get_legend()
            if leg: leg.remove()
    
    def set_ymax_callback(self, callback):
        """Set callback function to update ymax spinbox"""
        self._ymax_callback = callback
    
    def on_scroll(self, event):
        """Handle scroll wheel events to adjust ymax"""
        if event.inaxes != self.ax:
            return  # Only handle scroll when cursor is over the plot
        
        if event.button == 'up' or event.button == 'down':
            # Get current y limits
            ymin, ymax = self.ax.get_ylim()
            y_range = ymax - ymin
            
            # Calculate scroll step (5% of current range)
            step = y_range * 0.05
            
            # Adjust ymax based on scroll direction
            if event.button == 'up':
                new_ymax = ymax + step
            else:  # event.button == 'down'
                new_ymax = ymax - step
                # Prevent ymax from going below ymin
                if new_ymax <= ymin:
                    new_ymax = ymin + 0.01
            
            # Update ymax
            self._ymax = new_ymax
            self._ymin = ymin
            self._manual_y = True
            self._autoscale_y = False
            self.ax.set_ylim(ymin, new_ymax)
            self.draw_idle()
            
            # Update spinbox via callback
            if self._ymax_callback:
                self._ymax_callback(ymin, new_ymax)




class OceanMini(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("OceanMini — Ocean Insight Mini GUI")
        self.resize(1000, 600)
        

        central = QtWidgets.QWidget(self); self.setCentralWidget(central)
        layout = QtWidgets.QHBoxLayout(central)

        self.canvas = MplCanvas(self); layout.addWidget(self.canvas, stretch=4)

        panel = QtWidgets.QFrame(self); layout.addWidget(panel, stretch=0)
        form = QtWidgets.QFormLayout(panel)

        self.status_label = QtWidgets.QLabel("Status: —"); self.status_label.setWordWrap(True)
        form.addRow(self.status_label)

        row = QtWidgets.QHBoxLayout()
        self.btn_connect = QtWidgets.QPushButton("Connect")
        self.btn_start = QtWidgets.QPushButton("Start")
        self.btn_stop = QtWidgets.QPushButton("Stop")
        row.addWidget(self.btn_connect); row.addWidget(self.btn_start); row.addWidget(self.btn_stop)
        form.addRow(row)

        self.spin_it = QtWidgets.QSpinBox(); self.spin_it.setRange(1, 60_000); self.spin_it.setValue(200); self.spin_it.setSuffix(" ms")
        form.addRow("Integration Time:", self.spin_it)

        self.spin_avg = QtWidgets.QSpinBox(); self.spin_avg.setRange(1, 100); self.spin_avg.setValue(1)
        form.addRow("Averages:", self.spin_avg)

        self.chk_edark = QtWidgets.QCheckBox("Electrical dark correction")
        self.chk_nl = QtWidgets.QCheckBox("Nonlinearity correction")
        form.addRow(self.chk_edark); form.addRow(self.chk_nl)

        dark_row = QtWidgets.QHBoxLayout()
        self.btn_cap_dark = QtWidgets.QPushButton("Capture Dark")
        self.btn_clr_dark = QtWidgets.QPushButton("Clear Dark")
        dark_row.addWidget(self.btn_cap_dark); dark_row.addWidget(self.btn_clr_dark)
        form.addRow(dark_row)

        self.chk_apply_dark = QtWidgets.QCheckBox("Apply dark subtraction"); self.chk_apply_dark.setChecked(True)
        form.addRow(self.chk_apply_dark)
        
        # Auto-scale Y (so zoom isn’t overridden)
        self.chk_autoscale_y = QtWidgets.QCheckBox("Auto-scale Y (live)")
        self.chk_autoscale_y.setChecked(True)
        form.addRow(self.chk_autoscale_y)
        
        # Snapshots
        snap_row = QtWidgets.QHBoxLayout()
        self.btn_snap = QtWidgets.QPushButton("Take snapshot")
        self.btn_snap_clear = QtWidgets.QPushButton("Clear snapshots")
        snap_row.addWidget(self.btn_snap)
        snap_row.addWidget(self.btn_snap_clear)
        form.addRow(snap_row)
        
        self.chk_autoscale_y.toggled.connect(self.canvas.set_autoscale_y)
        self.btn_snap.clicked.connect(self.take_snapshot)
        self.btn_snap_clear.clicked.connect(self.clear_snapshots)
        
                # --- Manual Y controls ---
        y_row1 = QtWidgets.QHBoxLayout()
        self.spin_ymin = QtWidgets.QDoubleSpinBox()
        self.spin_ymax = QtWidgets.QDoubleSpinBox()
        for sp in (self.spin_ymin, self.spin_ymax):
            sp.setDecimals(3)
            sp.setRange(-1e9, 1e9)
            sp.setSingleStep(10.0)
        self.spin_ymin.setPrefix("Min: ")
        self.spin_ymax.setPrefix("Max: ")
        y_row1.addWidget(self.spin_ymin); y_row1.addWidget(self.spin_ymax)
        form.addRow(y_row1)
        
        y_row2 = QtWidgets.QHBoxLayout()
        self.btn_y_apply = QtWidgets.QPushButton("Apply Y limits")
        self.btn_y_reset = QtWidgets.QPushButton("Reset Y (auto)")
        y_row2.addWidget(self.btn_y_apply); y_row2.addWidget(self.btn_y_reset)
        form.addRow(y_row2)

        self.chk_autoscale_y.toggled.connect(self.canvas.set_autoscale_y)
        self.btn_y_apply.clicked.connect(self.on_apply_y_limits)
        self.btn_y_reset.clicked.connect(self.on_reset_y_limits)
        # Auto-apply Y limits when spinbox values change
        self.spin_ymin.valueChanged.connect(self.on_apply_y_limits)
        self.spin_ymax.valueChanged.connect(self.on_apply_y_limits)
        # Connect canvas scroll to update ymax spinbox
        self.canvas.set_ymax_callback(self.update_ymax_from_scroll)



        # Calibration data toggle
        self.chk_calibration = QtWidgets.QCheckBox("Calibration data")
        form.addRow(self.chk_calibration)

        # Calibration and pH prediction buttons
        cal_row = QtWidgets.QHBoxLayout()
        self.btn_calibrate = QtWidgets.QPushButton("Calibrate")
        self.btn_ph_predict = QtWidgets.QPushButton("pH Prediction")
        cal_row.addWidget(self.btn_calibrate)
        cal_row.addWidget(self.btn_ph_predict)
        form.addRow(cal_row)
        
        self.btn_calibrate.clicked.connect(self.run_calibration)
        self.btn_ph_predict.clicked.connect(self.predict_pH)

        self.btn_save = QtWidgets.QPushButton("Save CSV…")
        form.addRow(self.btn_save)

        # Regular Acquisition Controls
        form.addRow(QtWidgets.QLabel("--- Regular Acquisitions ---"))
        
        # Interval time control
        interval_row = QtWidgets.QHBoxLayout()
        self.spin_interval = QtWidgets.QSpinBox()
        self.spin_interval.setRange(1, 3600)  # 1 second to 1 hour
        self.spin_interval.setValue(30)  # Default 30 seconds
        self.spin_interval.setSuffix(" sec")
        interval_row.addWidget(QtWidgets.QLabel("Interval:"))
        interval_row.addWidget(self.spin_interval)
        form.addRow(interval_row)
        
        # Number of acquisitions control
        count_row = QtWidgets.QHBoxLayout()
        self.spin_count = QtWidgets.QSpinBox()
        self.spin_count.setRange(1, 1000)  # 1 to 1000 acquisitions
        self.spin_count.setValue(10)  # Default 10 acquisitions
        count_row.addWidget(QtWidgets.QLabel("Count:"))
        count_row.addWidget(self.spin_count)
        form.addRow(count_row)
        
        # Regular acquisitions filename control
        fname_row = QtWidgets.QHBoxLayout()
        self.le_reg_filename = QtWidgets.QLineEdit()
        self.le_reg_filename.setPlaceholderText("e.g. regular_acq")  # shown when empty
        self.le_reg_filename.setText("regular_acq")                  # default value
        fname_row.addWidget(QtWidgets.QLabel("Reg. acq file name:"))
        fname_row.addWidget(self.le_reg_filename)
        form.addRow(fname_row)

        
        # Regular acquisition buttons
        reg_acq_row = QtWidgets.QHBoxLayout()
        self.btn_start_reg = QtWidgets.QPushButton("Start Regular Acq")
        self.btn_stop_reg = QtWidgets.QPushButton("Stop Regular Acq")
        self.btn_stop_reg.setEnabled(False)
        reg_acq_row.addWidget(self.btn_start_reg)
        reg_acq_row.addWidget(self.btn_stop_reg)
        form.addRow(reg_acq_row)
        
        # Progress indicator
        self.progress_reg = QtWidgets.QProgressBar()
        self.progress_reg.setVisible(False)
        form.addRow("Progress:", self.progress_reg)
        
        # Status for regular acquisitions
        self.status_reg = QtWidgets.QLabel("Ready for regular acquisitions")
        self.status_reg.setStyleSheet("color: blue; font-size: 10px;")
        form.addRow(self.status_reg)

        self.footer = QtWidgets.QLabel(
            "Tips: Capture a dark with shutter closed, then tick 'Apply dark subtraction'. "
            "All device I/O is now serialized to prevent disconnects."
        )
        self.footer.setStyleSheet("color: gray; font-size: 11px;"); self.footer.setWordWrap(True)
        form.addRow(self.footer)

        self.thread = QtCore.QThread(self)
        self.worker = SpectrumWorker(); self.worker.moveToThread(self.thread); self.thread.start()

        self.worker.spectrumReady.connect(self.on_spectrum)
        self.worker.deviceStatus.connect(self.on_status)

        self.btn_connect.clicked.connect(self.worker.connect_device)
        self.btn_start.clicked.connect(self.worker.start)
        self.btn_stop.clicked.connect(self.worker.stop)

        self.spin_it.valueChanged.connect(self.worker.set_integration_ms)
        self.spin_avg.valueChanged.connect(self.worker.set_avg_window)
        self.chk_edark.toggled.connect(self.worker.set_correct_edark)
        self.chk_nl.toggled.connect(self.worker.set_correct_nonlinearity)
        self.chk_apply_dark.toggled.connect(self.worker.set_apply_dark_sub)
        self.btn_cap_dark.clicked.connect(self.worker.capture_dark)
        self.btn_clr_dark.clicked.connect(self.worker.clear_dark)
        self.btn_save.clicked.connect(self.save_csv)
        
        # Connect regular acquisition buttons
        self.btn_start_reg.clicked.connect(self.start_regular_acquisition)
        self.btn_stop_reg.clicked.connect(self.stop_regular_acquisition)
        
        self._wl_static = None
        self._last_int = None
        self._last_meta = None
        self._last_time = None
        
        # Calibration constants
        self._calib_Ia = None
        self._calib_Ib = None
        self._calib_Ka = None
        self._calib_valid = False
        
        # Regular acquisition state
        self._reg_acq_timer = None
        self._reg_acq_count = 0
        self._reg_acq_total = 0
        self._reg_acq_data = []  # Store acquisition data
        self._reg_acq_ph_series = []  # Store pH time series


        QtCore.QTimer.singleShot(300, self.worker.connect_device)

    @QtCore.pyqtSlot(np.ndarray, np.ndarray, float, dict)
    def on_spectrum(self, wl, inten, ts, meta):
        if self._wl_static is None:
            self._wl_static = wl  # store once; enables stable x for panning/zoom
    
        # live plot (does not accumulate)
        self.canvas.plot_live(self._wl_static, inten)
    
        t = datetime.fromtimestamp(ts).strftime('%H:%M:%S')
        self.statusBar().showMessage(
            f"{t} | IT={meta['integration_us']/1000:.0f} ms | avg={meta['averages']} | "
            f"eDark={'on' if meta['correct_edark'] else 'off'} | "
            f"NL={'on' if meta['correct_nonlinearity'] else 'off'} | "
            f"DarkSub={'on' if meta['apply_dark_sub'] else 'off'}"
        )
        # keep a reference for Snapshot/Save
        self._last_int = inten
        self._last_meta = meta
        self._last_time = ts
    def take_snapshot(self):
        if self._wl_static is None or self._last_int is None:
            QtWidgets.QMessageBox.information(self, "Snapshot", "No spectrum yet.")
            return
        label = datetime.fromtimestamp(self._last_time).strftime("Snap %H:%M:%S")
        # copy y so later changes don't mutate this snapshot
        self.canvas.add_snapshot(self._wl_static, self._last_int.copy(), label)
        
        # Save spectrum to CSV
        try:
            # Get current date and time
            now = datetime.now()
            date_folder = now.strftime('%Y-%m-%d')
            time_filename = now.strftime('%H-%M-%S')
            
            # Create date folder if it doesn't exist
            if not os.path.exists(date_folder):
                os.makedirs(date_folder)
            
            # Create CSV file path
            csv_filename = f"{time_filename}.csv"
            csv_path = os.path.join(date_folder, csv_filename)
            
            # Write CSV file
            with open(csv_path, "w", encoding="utf-8") as f:
                # Write header
                f.write("wavelength_nm,intensity_au\n")
                # Write data
                for wl, inten in zip(self._wl_static, self._last_int):
                    f.write(f"{wl:.6f},{inten:.6f}\n")
            
            self.statusBar().showMessage(f"Snapshot saved: {csv_path}", 5000)
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.warning(self, "Save Error", f"Failed to save snapshot CSV:\n{str(e)}")
    
    def clear_snapshots(self):
        self.canvas.clear_snapshots()
    def on_apply_y_limits(self):
        ymin = self.spin_ymin.value()
        ymax = self.spin_ymax.value()
        # turn off autoscale if user sets manual limits
        if self.chk_autoscale_y.isChecked():
            self.chk_autoscale_y.setChecked(False)
        self.canvas.set_manual_y(True, ymin, ymax)
    
    def update_ymax_from_scroll(self, ymin_value, ymax_value):
        """Update ymin and ymax spinboxes from scroll event (without triggering auto-apply)"""
        # Temporarily block signals to avoid triggering on_apply_y_limits
        self.spin_ymin.blockSignals(True)
        self.spin_ymax.blockSignals(True)
        self.spin_ymin.setValue(ymin_value)
        self.spin_ymax.setValue(ymax_value)
        self.spin_ymin.blockSignals(False)
        self.spin_ymax.blockSignals(False)
        # Also update autoscale checkbox state
        if self.chk_autoscale_y.isChecked():
            self.chk_autoscale_y.setChecked(False)
    
    def on_reset_y_limits(self):
        # return to autoscale (and also works nicely with user zoom if you untick autoscale after)
        self.canvas.reset_y()
        self.chk_autoscale_y.setChecked(True)


    @QtCore.pyqtSlot(str)
    def on_status(self, text):
        self.status_label.setText(f"Status: {text}")

    def save_csv(self):
        try:
            if self._wl_static is None:
                QtWidgets.QMessageBox.warning(self, "Save", "No spectrum to save yet."); return
            
            # Check if calibration mode is enabled
            if self.chk_calibration.isChecked():
                self.save_calibration_data()
            else:
                    # Original CSV save functionality
                default = f"spectrum_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Spectrum CSV", default, "CSV Files (*.csv)")
                if not path: return
                header_lines = [
                    "OceanMini CSV",
                    f"timestamp,{datetime.fromtimestamp(self._last_time).isoformat()}",
                    f"integration_time_us,{self._last_meta.get('integration_us','')}",
                    f"averages,{self._last_meta.get('averages','')}",
                    f"correct_electrical_dark,{self._last_meta.get('correct_edark','')}",
                    f"correct_nonlinearity,{self._last_meta.get('correct_nonlinearity','')}",
                    f"apply_dark_subtraction,{self._last_meta.get('apply_dark_sub','')}",
                    "wavelength_nm,intensity_au",
                ]
                with open(path, "w", encoding="utf-8") as f:
                    for line in header_lines: f.write(line + "\n")
                    for x, y in zip(self._wl_static, self._last_int):
                        f.write(f"{x:.6f},{y:.6f}\n")
                self.statusBar().showMessage(f"Saved: {path}", 5000)
        except Exception:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Save", "Failed to save.")

    def save_calibration_data(self):
        """Save calibration spectrum to meas_data.xlsx"""
        try:
            # Ask for pH value
            pH_value, ok = QtWidgets.QInputDialog.getDouble(
                self, 
                "pH Value", 
                "Enter pH value:", 
                value=7.0,  # default value
                min=0.0,
                max=14.0,
                decimals=2
            )
            
            if not ok:
                return  # User cancelled
            
            # Excel file path (in current working directory)
            excel_path = "meas_data.xlsx"
            
            # Load or create workbook
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                if "calibration" in wb.sheetnames:
                    ws = wb["calibration"]
                else:
                    ws = wb.create_sheet("calibration")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "calibration"
            
            # Find the next available column (first empty column)
            next_col = 1
            while ws.cell(row=1, column=next_col).value is not None:
                next_col += 2  # Skip pairs of columns (wavelength + intensity)
            
            # Write headers
            ws.cell(row=1, column=next_col, value=f"pH {pH_value}")
            ws.cell(row=1, column=next_col + 1, value="")  # Empty title for intensity
            
            # Write data
            for i, (wl, inten) in enumerate(zip(self._wl_static, self._last_int), start=2):
                ws.cell(row=i, column=next_col, value=float(wl))
                ws.cell(row=i, column=next_col + 1, value=float(inten))
            
            # Save workbook
            wb.save(excel_path)
            
            self.statusBar().showMessage(f"Calibration data saved to {excel_path} (pH {pH_value})", 5000)
            QtWidgets.QMessageBox.information(
                self, 
                "Calibration Saved", 
                f"Calibration data saved successfully!\nFile: {excel_path}\npH: {pH_value}"
            )
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Save Calibration", f"Failed to save calibration data:\n{str(e)}")

    def run_calibration(self):
        """Load calibration data from Excel file and fit the model"""
        try:
            # Excel file path
            excel_path = "meas_data.xlsx"
            
            if not os.path.exists(excel_path):
                QtWidgets.QMessageBox.warning(
                    self, 
                    "Calibration Error", 
                    f"Calibration file '{excel_path}' not found.\nPlease save calibration data first."
                )
                return
            
            # Load calibration sheet
            df = pd.read_excel(excel_path, sheet_name="calibration")
            
            # Parse the pH-wavelength-intensity pairs
            pairs = parse_xy_pairs(df)
            
            if not pairs:
                QtWidgets.QMessageBox.warning(
                    self, 
                    "Calibration Error", 
                    "No valid calibration spectra found in the Excel file."
                )
                return
            
            # Process each spectrum: find left peak center and get filtered intensity
            records = []
            for p in pairs:
                x, y = p["x"], p["y"]
                try:
                    lam1 = fit_left_center_lambda(x, y)
                    y_f = savgol_filter(y, auto_win(len(y)), SMOOTH_POLY) if len(y) >= 7 else y
                    I_at = float(np.interp(lam1, x, y_f))
                    records.append({"pH": p["pH"], "lambda1": lam1, "I": I_at})
                except Exception as e:
                    print(f"Warning: Failed to process spectrum at pH {p['pH']}: {e}")
                    continue
            
            if len(records) < 3:
                QtWidgets.QMessageBox.warning(
                    self, 
                    "Calibration Error", 
                    f"Not enough valid spectra for calibration (found {len(records)}, need at least 3)."
                )
                return
            
            # Create dataframe and sort by pH
            data = pd.DataFrame(records).sort_values('pH').reset_index(drop=True)
            
            # Fit the model
            pH_arr = data['pH'].values
            I_arr = data['I'].values
            
            Ia, Ib, Ka = fit_Ia_Ib_Ka(pH_arr, I_arr)
            pKa = -np.log10(Ka)
            
            # Store calibration constants
            self._calib_Ia = Ia
            self._calib_Ib = Ib
            self._calib_Ka = Ka
            self._calib_valid = True
            
            # Show results
            QtWidgets.QMessageBox.information(
                self, 
                "Calibration Complete", 
                f"Calibration successful!\n\n"
                f"Fitted constants:\n"
                f"  Ia = {Ia:.3f}\n"
                f"  Ib = {Ib:.3f}\n"
                f"  Ka = {Ka:.3e}\n"
                f"  pKa = {pKa:.3f}\n\n"
                f"Used {len(records)} calibration spectra\n"
                f"pH range: {pH_arr.min():.2f} - {pH_arr.max():.2f}"
            )
            
            self.statusBar().showMessage(
                f"Calibration complete: Ia={Ia:.1f}, Ib={Ib:.1f}, pKa={pKa:.2f}", 
                10000
            )
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "Calibration Error", 
                f"Failed to perform calibration:\n{str(e)}"
            )

    def predict_pH(self):
        """Predict pH from current spectrum and optionally save it"""
        try:
            if self._wl_static is None or self._last_int is None:
                QtWidgets.QMessageBox.warning(self, "pH Prediction", "No spectrum available.")
                return
            
            if not self._calib_valid:
                QtWidgets.QMessageBox.warning(
                    self, 
                    "pH Prediction", 
                    "No calibration data loaded.\nPlease run calibration first."
                )
                return
            
            # Process current spectrum
            x, y = self._wl_static, self._last_int
            
            # Find left peak center
            lam1 = fit_left_center_lambda(x, y)
            
            # Apply Savitzky-Golay filter
            y_f = savgol_filter(y, auto_win(len(y)), SMOOTH_POLY) if len(y) >= 7 else y
            
            # Get intensity at left peak center
            I_at = float(np.interp(lam1, x, y_f))
            
            # Predict pH
            pH_pred = invert_to_pH(I_at, self._calib_Ia, self._calib_Ib, self._calib_Ka)
            
            if not np.isfinite(pH_pred):
                QtWidgets.QMessageBox.warning(
                    self, 
                    "pH Prediction", 
                    f"Unable to predict pH.\nIntensity value ({I_at:.1f}) may be outside calibration range."
                )
                return
            
            # Show prediction
            msg = QtWidgets.QMessageBox(self)
            msg.setWindowTitle("pH Prediction")
            msg.setText(
                f"Predicted pH: {pH_pred:.2f}\n\n"
                f"Left peak wavelength: {lam1:.2f} nm\n"
                f"Intensity at peak: {I_at:.1f}\n\n"
                f"Do you want to save this spectrum?"
            )
            msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            msg.setDefaultButton(QtWidgets.QMessageBox.Yes)
            
            result = msg.exec_()
            
            if result == QtWidgets.QMessageBox.Yes:
                self.save_prediction_spectrum(pH_pred, lam1, I_at)
                
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "pH Prediction Error", 
                f"Failed to predict pH:\n{str(e)}"
            )

    def save_prediction_spectrum(self, pH_pred, lambda1, intensity):
        """Save prediction spectrum to Excel with user-selected sheet"""
        try:
            excel_path = "meas_data.xlsx"
            
            # Load or create workbook
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                existing_sheets = wb.sheetnames
            else:
                wb = Workbook()
                existing_sheets = []
            
            # Ask user for sheet name
            sheet_name, ok = QtWidgets.QInputDialog.getItem(
                self,
                "Select Sheet",
                "Select or type sheet name for prediction data:",
                existing_sheets + ["<New Sheet>"],
                0,
                True  # Allow editing
            )
            
            if not ok or not sheet_name:
                return
            
            # Handle new sheet or validate name
            if sheet_name == "<New Sheet>" or sheet_name.strip() == "":
                sheet_name, ok = QtWidgets.QInputDialog.getText(
                    self,
                    "New Sheet Name",
                    "Enter name for new sheet:"
                )
                if not ok or not sheet_name.strip():
                    return
                sheet_name = sheet_name.strip()
            
            # Get or create the sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Find the next available column
            next_col = 1
            while ws.cell(row=1, column=next_col).value is not None:
                next_col += 2
            
            # Create timestamp label
            timestamp = datetime.fromtimestamp(self._last_time).strftime('%Y-%m-%d %H:%M:%S')
            
            # Write headers
            ws.cell(row=1, column=next_col, value=f"pH {pH_pred:.2f} @ {timestamp}")
            ws.cell(row=1, column=next_col + 1, value="")
            
            # Write data
            for i, (wl, inten) in enumerate(zip(self._wl_static, self._last_int), start=2):
                ws.cell(row=i, column=next_col, value=float(wl))
                ws.cell(row=i, column=next_col + 1, value=float(inten))
            
            # Save workbook
            wb.save(excel_path)
            
            self.statusBar().showMessage(
                f"Prediction saved to '{sheet_name}' sheet (pH {pH_pred:.2f})", 
                5000
            )
            QtWidgets.QMessageBox.information(
                self,
                "Saved",
                f"Spectrum saved successfully!\n\n"
                f"File: {excel_path}\n"
                f"Sheet: {sheet_name}\n"
                f"Predicted pH: {pH_pred:.2f}"
            )
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "Save Error", 
                f"Failed to save prediction spectrum:\n{str(e)}"
            )

    def start_regular_acquisition(self):
        """Start regular acquisition sequence"""
        try:
            # if not self._calib_valid:
            #     QtWidgets.QMessageBox.warning(
            #         self, 
            #         "Regular Acquisition", 
            #         "No calibration data loaded.\nPlease run calibration first."
            #     )
            #     return
            
            if self._wl_static is None:
                QtWidgets.QMessageBox.warning(
                    self, 
                    "Regular Acquisition", 
                    "No spectrum available.\nPlease start streaming first."
                )
                return
            
            # Get parameters
            interval_sec = self.spin_interval.value()
            total_count = self.spin_count.value()
            
            # Reset state
            self._reg_acq_count = 0
            self._reg_acq_total = total_count
            self._reg_acq_data = []
            self._reg_acq_ph_series = []
            
            # Update UI
            self.btn_start_reg.setEnabled(False)
            self.btn_stop_reg.setEnabled(True)
            self.progress_reg.setVisible(True)
            self.progress_reg.setMaximum(total_count)
            self.progress_reg.setValue(0)
            self.status_reg.setText(f"Starting regular acquisition: {total_count} acquisitions every {interval_sec}s")
            
            # Start timer
            self._reg_acq_timer = QtCore.QTimer()
            self._reg_acq_timer.timeout.connect(self.perform_regular_acquisition)
            self._reg_acq_timer.start(interval_sec * 1000)  # Convert to milliseconds
            
            # Perform first acquisition immediately
            self.perform_regular_acquisition()
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "Regular Acquisition Error", 
                f"Failed to start regular acquisition:\n{str(e)}"
            )

    def perform_regular_acquisition(self):
        """Perform a single acquisition in the regular sequence"""
        try:
            if self._reg_acq_count >= self._reg_acq_total:
                self.finish_regular_acquisition()
                return
            
            if self._wl_static is None or self._last_int is None:
                self.status_reg.setText("No spectrum available for acquisition")
                return
            
            # Process current spectrum for pH prediction
            x, y = self._wl_static, self._last_int
            
            # # Find left peak center
            # lam1 = fit_left_center_lambda(x, y)
            
            # # Apply Savitzky-Golay filter
            # y_f = savgol_filter(y, auto_win(len(y)), SMOOTH_POLY) if len(y) >= 7 else y
            
            # # Get intensity at left peak center
            # I_at = float(np.interp(lam1, x, y_f))
            
            # # Predict pH
            # pH_pred = invert_to_pH(I_at, self._calib_Ia, self._calib_Ib, self._calib_Ka)
            
            # if not np.isfinite(pH_pred):
            #     self.status_reg.setText(f"Acquisition {self._reg_acq_count + 1}: Invalid pH prediction")
            #     self._reg_acq_count += 1
            #     self.progress_reg.setValue(self._reg_acq_count)
            #     return
            
            # Store acquisition data
            timestamp = datetime.fromtimestamp(self._last_time)
            acquisition_data = {
                'order': self._reg_acq_count + 1,
                'timestamp': timestamp,
                # 'pH_pred': pH_pred,
                # 'wavelength': lam1,
                # 'intensity': I_at,
                'spectrum_wl': x.copy(),
                'spectrum_int': y.copy()
            }
            
            self._reg_acq_data.append(acquisition_data)
            self._reg_acq_ph_series.append({
                'timestamp': timestamp,
                # 'pH': pH_pred
            })
            
            # Update UI
            self._reg_acq_count += 1
            self.progress_reg.setValue(self._reg_acq_count)
            self.status_reg.setText(
                f"Acquisition {self._reg_acq_count}/{self._reg_acq_total}: "
                # f"pH = {pH_pred:.2f} (λ = {lam1:.1f} nm)"
            )
            
            # Check if we're done
            if self._reg_acq_count >= self._reg_acq_total:
                self.finish_regular_acquisition()
                
        except Exception as e:
            traceback.print_exc()
            self.status_reg.setText(f"Error in acquisition {self._reg_acq_count + 1}: {str(e)}")
            self._reg_acq_count += 1
            self.progress_reg.setValue(self._reg_acq_count)

    def finish_regular_acquisition(self):
        """Finish the regular acquisition sequence and save data"""
        try:
            # Stop timer
            if self._reg_acq_timer:
                self._reg_acq_timer.stop()
                self._reg_acq_timer = None
            
            # Update UI
            self.btn_start_reg.setEnabled(True)
            self.btn_stop_reg.setEnabled(False)
            self.status_reg.setText(f"Regular acquisition complete: {self._reg_acq_count} acquisitions")
            
            # Save data
            self.save_regular_acquisition_data()
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "Regular Acquisition Error", 
                f"Failed to finish regular acquisition:\n{str(e)}"
            )

    def stop_regular_acquisition(self):
        """Stop the regular acquisition sequence"""
        try:
            # Stop timer
            if self._reg_acq_timer:
                self._reg_acq_timer.stop()
                self._reg_acq_timer = None
            
            # Update UI
            self.btn_start_reg.setEnabled(True)
            self.btn_stop_reg.setEnabled(False)
            self.status_reg.setText(f"Regular acquisition stopped: {self._reg_acq_count} acquisitions completed")
            
            # Save partial data if any
            if self._reg_acq_count > 0:
                self.save_regular_acquisition_data()
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "Regular Acquisition Error", 
                f"Failed to stop regular acquisition:\n{str(e)}"
            )

    def save_regular_acquisition_data(self):
        """Save regular acquisition data to Excel files"""
        try:
            if not self._reg_acq_data:
                return
            
            # Save main acquisition data (all spectra in one sheet)
            self.save_acquisition_spectra()
            
            # Save pH time series
            # self.save_ph_time_series()
            
            # Show completion message
            QtWidgets.QMessageBox.information(
                self,
                "Regular Acquisition Complete",
                f"Regular acquisition completed successfully!\n\n"
                f"Total acquisitions: {len(self._reg_acq_data)}\n"
                f"Data saved to Excel files."
            )
            
        except Exception as e:
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, 
                "Save Error", 
                f"Failed to save regular acquisition data:\n{str(e)}"
            )

    def save_acquisition_spectra(self):
        """Save all acquisition spectra to one Excel sheet"""
        try:
            now = datetime.now()
            date_folder = now.strftime('%Y-%m-%d')
            base_name = (self.le_reg_filename.text() or "").strip()
            if not base_name:
                base_name = "regular_acq"
                
            time_filename = now.strftime('%H-%M-%S')
            excel_filename = f"{base_name}_{time_filename}.csv"
            
            # Create date folder if it doesn't exist
            if not os.path.exists(date_folder):
                os.makedirs(date_folder)

            excel_path = os.path.join(date_folder, excel_filename)
            
            # Load or create workbook
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
            
            # Create or get regular acquisitions sheet
            sheet_name = "regular_acquisitions"
            ws = wb.active          # get default sheet
            ws.title = sheet_name   # rename it
            # if sheet_name in wb.sheetnames:
            #     ws = wb[sheet_name]
            #     # Clear existing data
            #     ws.delete_rows(1, ws.max_row)
            # else:
            #     ws = wb.create_sheet(sheet_name)
            
            # Write data for each acquisition
            for i, acq in enumerate(self._reg_acq_data):
                col_start = i * 2 + 1
                
                # Create title with predicted pH, order number, and time
                # title = f"pH {acq['pH_pred']:.2f} - Acq #{acq['order']} - {acq['timestamp'].strftime('%H:%M:%S')}"
                title = f"Acq #{acq['order']} - {acq['timestamp'].strftime('%H:%M:%S')}"
                ws.cell(row=1, column=col_start, value=title)
                ws.cell(row=1, column=col_start + 1, value="")  # Empty title for intensity
                
                # Write spectrum data
                for j, (wl, inten) in enumerate(zip(acq['spectrum_wl'], acq['spectrum_int']), start=2):
                    ws.cell(row=j, column=col_start, value=float(wl))
                    ws.cell(row=j, column=col_start + 1, value=float(inten))
            
            # Save workbook
            wb.save(excel_path)
            
            self.statusBar().showMessage(f"Acquisition spectra saved to {excel_path}", 5000)
            
        except Exception as e:
            raise Exception(f"Failed to save acquisition spectra: {str(e)}")

    def save_ph_time_series(self):
        """Save pH time series to separate Excel file"""
        try:
            if not self._reg_acq_ph_series:
                return
            
            # Create time series file
            ts_file = "ph_time_series.xlsx"
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "pH_TimeSeries"
            
            # Write headers
            ws.cell(row=1, column=1, value="Timestamp")
            ws.cell(row=1, column=2, value="pH")
            ws.cell(row=1, column=3, value="Order")
            
            # Write data
            for i, data in enumerate(self._reg_acq_ph_series, start=2):
                ws.cell(row=i, column=1, value=data['timestamp'])
                ws.cell(row=i, column=2, value=data['pH'])
                ws.cell(row=i, column=3, value=i-1)  # Order number
            
            # Save workbook
            wb.save(ts_file)
            
            self.statusBar().showMessage(f"pH time series saved to {ts_file}", 5000)
            
        except Exception as e:
            raise Exception(f"Failed to save pH time series: {str(e)}")

    def closeEvent(self, event):
        try:
            self.worker.stop()
            self.thread.quit(); self.thread.wait(1500)
        except Exception:
            pass
        super().closeEvent(event)


def main():
    app = QtWidgets.QApplication(sys.argv)
    w = OceanMini(); w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
